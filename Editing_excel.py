import pandas as pd 
import numpy as np


df = pd.read_excel (r'G:\study\internshipp\CSIR_CDRI\Brown_University.xlsx')

import xlrd 

loc = (r'G:\study\internshipp\CSIR_CDRI\Shelton_commmunity_college.xlsx') 
  
# To open Workbook 
wb = xlrd.open_workbook(loc) 
sheet = wb.sheet_by_index(0) 
sheet.getCells()
sheet.cell_value(2,0)


import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
font = Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000')
wb = openpyxl.load_workbook(r'G:\study\internshipp\CSIR_CDRI\Brown_University.xlsx') 
sheet = wb.active 


for i in range(50,len(wb.sheetnames)):
    wb.worksheets[i].delete_cols(1)
    wb.worksheets[i].insert_rows(1)
    wb.worksheets[i].insert_rows(1)
    wb.worksheets[i].insert_rows(1)
    wb.worksheets[i].insert_rows(1)
    wb.worksheets[i].cell(row = 1, column = 2).value=wb.sheetnames[i]
    wb.worksheets[i].cell(row = 1, column = 2).font=font
    wb.save(filename=r'G:\study\internshipp\CSIR_CDRI\Brown_University.xlsx')
