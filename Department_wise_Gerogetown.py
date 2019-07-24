from urllib.request import urlopen
from bs4 import BeautifulSoup
from selenium import webdriver
import time
import pandas as pd
import numpy as np
from openpyxl import load_workbook
url = "http://www.georgetowncollege.edu/academics/faculty-directory?tid=All&keys="
driver = webdriver.PhantomJS(r"G:\study\software\phantomjs-2.1.1-windows\phantomjs-2.1.1-windows\bin\phantomjs.exe")
driver.get(url)

# This will get the initial html - before javascript
html1 = driver.page_source
soup = BeautifulSoup(html1)


categories=soup.find_all('option')
categories=categories[1:]
categories=categories[:-5]
categories
categories[0].text
id=[]
for i in range(0,len(categories)):
    id.append(str(categories[i])[15:17])

url1="http://www.georgetowncollege.edu/academics/faculty-directory?tid="+id[0]+"&keys="
driver.get(url1)
html1 = driver.page_source
soup = BeautifulSoup(html1)

email=soup.find_all('h5',attrs={'class':"margin-top: 5px;"})
str(email[2]).split('>')[1][:-1][16:][2:]

name=soup.find_all('h4',attrs={'class':"media-heading text-orange"})
name[0].text
education = soup.find_all('em')
education[2].text[2:]


book = load_workbook(r'G:\study\internshipp\CSIR_CDRI\Georgetown College.xlsx')
writer = pd.ExcelWriter(r'G:\study\internshipp\CSIR_CDRI\Georgetown College.xlsx', engine='openpyxl') 
writer.book = book

for i in range(0,len(categories)):
    print(i)
    categorie_name=categories[i].text
    url1="http://www.georgetowncollege.edu/academics/faculty-directory?tid="+id[i]+"&keys="
    driver.get(url1)
    time.sleep(2)
    html1 = driver.page_source
    soup = BeautifulSoup(html1)
    name1=[]
    email1=[]
    education1=[]
    name=soup.find_all('h4',attrs={'class':"media-heading text-orange"})
    email=soup.find_all('h5',attrs={'class':"margin-top: 5px;"})
    education = soup.find_all('em')
    for j in range(0,len(name)):
        name1.append(name[j].text)
        email1.append(str(email[j]).split('>')[1][:-1][16:][2:])
        education1.append(education[j].text[2:])
    z=pd.DataFrame(list(zip(name1,email1,education1)),columns=['Name','Email','Education'])
    book = load_workbook(r'G:\study\internshipp\CSIR_CDRI\Georgetown College.xlsx')
    writer = pd.ExcelWriter(r'G:\study\internshipp\CSIR_CDRI\Georgetown College.xlsx', engine='openpyxl') 
    writer.book = book
    z.to_excel(writer, sheet_name =categorie_name)        
    writer.save()











